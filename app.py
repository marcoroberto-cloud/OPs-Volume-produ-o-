import io
import os
from datetime import datetime
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
import pandas as pd
import streamlit as st

st.set_page_config(
    page_title="Gestão de OPs | Produção",
    page_icon="🏭",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# --- CSS AJUSTADO PARA NÃO TAMPAR O TOPO E MANTER A INTERFACE JUSTA ---
st.markdown(
    """
    <style>
        .block-container {
            padding-top: 3.2rem !important;
            padding-bottom: 1.5rem !important;
            padding-left: 1.5rem !important;
            padding-right: 1.5rem !important;
        }
        div[data-testid="stMetric"] {
            background-color: #161b22;
            padding: 8px 12px;
            border-radius: 8px;
            border: 1px solid #30363d;
        }
        div[data-testid="stMetricValue"] {
            font-size: 1.35rem !important;
            font-weight: 700;
        }
        div[data-testid="stMetricLabel"] {
            font-size: 0.8rem !important;
            color: #8b949e;
        }
        hr {
            margin: 0.6rem 0 !important;
        }
    </style>
    """,
    unsafe_allow_html=True,
)

DATA_CACHE_PATH = "base_ops_filtrada.parquet"
INFO_CACHE_PATH = "info_carga_ops.txt"


# --- PROCESSAMENTO EXATO DA MACRO VBA V11 ---
@st.cache_data(show_spinner=False)
def processar_arquivo_op(file_bytes):
    encodings = ["latin-1", "iso-8859-1", "cp1252", "utf-8"]
    df = None

    for enc in encodings:
        try:
            stream = io.BytesIO(file_bytes)
            preview = pd.read_csv(
                stream, sep=";", encoding=enc, nrows=5, header=None
            )
            skip = 0
            for idx, row in preview.iterrows():
                linha = " ".join([str(v) for v in row.values])
                if any(
                    k in linha
                    for k in ["Numero da OP", "Filial", "Produto", "SC2"]
                ):
                    skip = idx if "SC2" not in linha else 2
                    break

            stream.seek(0)
            df = pd.read_csv(
                stream,
                sep=";",
                skiprows=skip,
                encoding=enc,
                low_memory=False,
                dtype=str,
            )
            break
        except Exception:
            continue

    if df is None or df.empty:
        raise ValueError("Não foi possível ler o arquivo CSV do sistema.")

    df.columns = [str(c).strip() for c in df.columns]

    # Critérios da Macro: Filial 0101, Produto 46*00 e Sem PRIA
    filial_str = (
        df["Filial"].astype(str).str.strip().str.zfill(4)
        if "Filial" in df.columns
        else pd.Series("", index=df.index)
    )
    prod_str = (
        df["Produto"].astype(str).str.strip()
        if "Produto" in df.columns
        else pd.Series("", index=df.index)
    )

    cond_filial = (filial_str == "0101") | (
        df["Filial"].astype(str).str.strip() == "101"
    )
    cond_produto = prod_str.str.startswith("46") & prod_str.str.endswith("00")
    cond_not_pria = ~prod_str.str.upper().str.contains("PRIA", na=False)

    df_filtrado = df[cond_filial & cond_produto & cond_not_pria].copy()

    for col in ["Quantidade", "Qtd.Produzid"]:
        if col in df_filtrado.columns:
            df_filtrado[col] = (
                pd.to_numeric(
                    df_filtrado[col]
                    .astype(str)
                    .str.replace(".", "", regex=False)
                    .str.replace(",", ".", regex=False)
                    .str.strip(),
                    errors="coerce",
                )
                .fillna(0)
                .astype(float)
            )

    df_filtrado["Saldo_Pendente"] = (
        df_filtrado["Quantidade"] - df_filtrado["Qtd.Produzid"]
    ).clip(lower=0)

    if "DT Real Fim" in df_filtrado.columns:
        df_filtrado["DT_Real_Fim_Parsed"] = pd.to_datetime(
            df_filtrado["DT Real Fim"], format="%d/%m/%Y", errors="coerce"
        )
        df_filtrado["Ano"] = df_filtrado["DT_Real_Fim_Parsed"].dt.strftime("%Y").fillna("Sem Data")
        df_filtrado["Mes"] = df_filtrado["DT_Real_Fim_Parsed"].dt.strftime("%m").fillna("Sem Data")
        df_filtrado["MÊS-ANO"] = df_filtrado["DT_Real_Fim_Parsed"].dt.strftime("%Y/%m").fillna("Sem Data")
    else:
        df_filtrado["Ano"] = "Sem Data"
        df_filtrado["Mes"] = "Sem Data"
        df_filtrado["MÊS-ANO"] = "Sem Data"

    df_filtrado["STATUS"] = df_filtrado.apply(
        lambda r: "Ok"
        if (r["Quantidade"] == r["Qtd.Produzid"] and r["Quantidade"] > 0)
        else "Falta",
        axis=1,
    )

    for col_txt in ["Observacao", "Desc. Prod.", "Produto", "Numero da OP"]:
        if col_txt in df_filtrado.columns:
            df_filtrado[col_txt] = (
                df_filtrado[col_txt].fillna("").astype(str).str.strip()
            )

    return df_filtrado


def carregar_base_salva():
    if os.path.exists(DATA_CACHE_PATH) and os.path.exists(INFO_CACHE_PATH):
        try:
            df = pd.read_parquet(DATA_CACHE_PATH)
            with open(INFO_CACHE_PATH, "r", encoding="utf-8") as f:
                info_carga = f.read().strip()
            return df, info_carga
        except Exception:
            return None, "Nenhum arquivo salvo ainda"
    return None, "Nenhum arquivo salvo ainda"


def salvar_base_em_disco(df):
    df.to_parquet(DATA_CACHE_PATH, index=False)
    timestamp = datetime.now().strftime("%d/%m/%Y às %H:%M")
    with open(INFO_CACHE_PATH, "w", encoding="utf-8") as f:
        f.write(timestamp)
    return timestamp


@st.cache_data(show_spinner=False)
def gerar_resumo_mensal(df):
    resumo = (
        df.groupby("MÊS-ANO")
        .agg(
            Qtd_Planejada=("Quantidade", "sum"),
            Total_Produzido=("Qtd.Produzid", "sum"),
            Saldo_Pendente=("Saldo_Pendente", "sum"),
            Total_OPs=("Numero da OP", "count"),
        )
        .reset_index()
        .rename(columns={"Total_Produzido": "Total Produzido"})
    )
    return resumo.sort_values(by="MÊS-ANO", ascending=False)


@st.cache_data(show_spinner=False)
def gerar_excel_vba(df, resumo):
    buffer = io.BytesIO()
    nome_aba_dados = f"Extracao_{datetime.now().strftime('%d-%m_%H%M')}"
    cols_export = [
        "Filial",
        "Observacao",
        "Numero da OP",
        "Produto",
        "Desc. Prod.",
        "Quantidade",
        "Qtd.Produzid",
        "DT Real Fim",
        "STATUS",
        "MÊS-ANO",
        "DT Emissao",
    ]
    cols_existentes = [c for c in cols_export if c in df.columns]
    df_export = df[cols_existentes].copy()

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df_export.to_excel(
            writer, sheet_name=nome_aba_dados, index=False, startrow=9
        )
        resumo.to_excel(
            writer, sheet_name="RESUMO_MENSAL", index=False, startrow=2
        )
        wb = writer.book

        ws_dados = wb[nome_aba_dados]
        ws_dados.views.sheetView[0].showGridLines = True
        header_fill = PatternFill(
            start_color="1F4E78", end_color="1F4E78", fill_type="solid"
        )
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

        for cell in ws_dados[10]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for col in ws_dados.columns:
            max_len = max(len(str(cell.value or "")) for cell in col[9:80])
            col_letter = get_column_letter(col[0].column)
            ws_dados.column_dimensions[col_letter].width = max(max_len + 3, 12)

        ws_resumo = wb["RESUMO_MENSAL"]
        ws_resumo.views.sheetView[0].showGridLines = True
        ws_resumo.cell(row=1, column=1, value="RESUMO PRODUÇÃO MENSAL").font = (
            Font(name="Calibri", size=14, bold=True, color="1F4E78")
        )

        for cell in ws_resumo[3]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for col in ws_resumo.columns:
            max_len = max(len(str(cell.value or "")) for cell in col[2:30])
            col_letter = get_column_letter(col[0].column)
            ws_resumo.column_dimensions[col_letter].width = max(max_len + 4, 15)

    buffer.seek(0)
    return buffer


# --- INICIALIZAÇÃO ---
if "df_ops" not in st.session_state:
    df_salvo, info_salva = carregar_base_salva()
    st.session_state["df_ops"] = df_salvo
    st.session_state["info_carga"] = info_salva

# --- CABEÇALHO SUPERIOR ---
col_logo, col_upload, col_reset, col_mob = st.columns([1.8, 2.2, 0.6, 0.8])

with col_logo:
    st.markdown("### 🏭 Gestão de OPs")
    st.caption(f"🕒 Carga: {st.session_state['info_carga']}")

with col_upload:
    arquivo = st.file_uploader(
        "Carregar planilha de OPs (.csv)",
        type=["csv", "xlsx", "xls"],
        label_visibility="collapsed",
        key="uploader_ops_topo",
    )
    if arquivo is not None:
        file_bytes = arquivo.read()
        df_novo = processar_arquivo_op(file_bytes)
        info_nova = salvar_base_em_disco(df_novo)
        st.session_state["df_ops"] = df_novo
        st.session_state["info_carga"] = info_nova
        st.rerun()

with col_reset:
    if st.button("🧹 Resetar", use_container_width=True):
        if os.path.exists(DATA_CACHE_PATH):
            os.remove(DATA_CACHE_PATH)
        if os.path.exists(INFO_CACHE_PATH):
            os.remove(INFO_CACHE_PATH)
        st.session_state["df_ops"] = None
        st.session_state["info_carga"] = "Nenhum arquivo salvo ainda"
        st.rerun()

with col_mob:
    modo_mobile = st.toggle("📱 Modo Celular", value=False)

df_base = st.session_state["df_ops"]

if df_base is None or df_base.empty:
    st.info(
        "👆 Selecione ou arraste o arquivo CSV no campo acima para carregar o painel."
    )
    st.stop()

# --- FILTROS NO TOPO ---
col_busca_proj, col_busca_peca, col_filtro_status = st.columns([2.0, 1.2, 0.9])

with col_busca_proj:
    obs_disponiveis = sorted(
        [o for o in df_base["Observacao"].unique() if o.strip() != ""]
    )
    busca_projeto = st.multiselect(
        "📄 Digite e Flegue o(s) Lote(s) / Observação / Projeto:",
        options=obs_disponiveis,
        placeholder="Digite parte do nome ou código (Ex: TAT, RANGER, COROLLA)...",
    )

with col_busca_peca:
    busca_peca = st.text_input(
        "🔍 Buscar Peça:",
        placeholder="Código ou descrição...",
    )

with col_filtro_status:
    filtro_status_btn = st.radio(
        "📌 Status:",
        options=["Todos", "Falta", "Ok"],
        horizontal=True,
    )

# --- FILTRO EM DUAS ETAPAS: ANO E MÊS ---
col_ano, col_mes = st.columns([1, 2])

with col_ano:
    anos_disponiveis = sorted(
        [a for a in df_base["Ano"].unique() if a != "Sem Data"], reverse=True
    )
    if "Sem Data" in df_base["Ano"].unique():
        anos_disponiveis.append("Sem Data")
        
    anos_selecionados = st.multiselect(
        "📅 Filtrar Ano:",
        options=anos_disponiveis,
        placeholder="Todos os anos",
    )

with col_mes:
    # Se selecionou ano, traz apenas os meses daquele ano
    if anos_selecionados:
        meses_disp = sorted(
            df_base[df_base["Ano"].isin(anos_selecionados)]["MÊS-ANO"].unique().tolist(),
            reverse=True,
        )
    else:
        meses_disp = sorted(df_base["MÊS-ANO"].unique().tolist(), reverse=True)

    meses_selecionados = st.multiselect(
        "🗓️ Filtrar Mês:",
        options=meses_disp,
        placeholder="Todos os meses",
    )

# --- APLICAÇÃO DOS FILTROS (SE NÃO SELECIONAR NADA, TRAZ TODOS) ---
df_view = df_base.copy()

if busca_projeto:
    df_view = df_view[df_view["Observacao"].isin(busca_projeto)]

if busca_peca.strip():
    termo = busca_peca.strip().lower()
    df_view = df_view[
        df_view["Produto"].str.lower().str.contains(termo)
        | df_view["Desc. Prod."].str.lower().str.contains(termo)
    ]

if filtro_status_btn != "Todos":
    df_view = df_view[df_view["STATUS"] == filtro_status_btn]

if meses_selecionados:
    df_view = df_view[df_view["MÊS-ANO"].isin(meses_selecionados)]
elif anos_selecionados:
    df_view = df_view[df_view["Ano"].isin(anos_selecionados)]

# --- CÁLCULO DOS INDICADORES ---
qtd_total_prog = int(df_view["Quantidade"].sum())
qtd_total_prod = int(df_view["Qtd.Produzid"].sum())
qtd_total_pend = int(df_view["Saldo_Pendente"].sum())
perc_ating = (
    (qtd_total_prod / qtd_total_prog * 100) if qtd_total_prog > 0 else 0.0
)

# --- CARDS DE KPI COMPACTOS ---
kpi1, kpi2, kpi3, kpi4 = st.columns(4)

kpi1.metric(
    "1. Programado (OP)",
    f"{qtd_total_prog:,.0f} pçs".replace(",", "."),
    f"Total OPs: {len(df_view):,}".replace(",", "."),
    delta_color="off",
)
kpi2.metric(
    "2. Fabricado / Produzido",
    f"{qtd_total_prod:,.0f} pçs".replace(",", "."),
    f"{perc_ating:.1f}% Concluído",
    delta_color="normal",
)
kpi3.metric(
    "3. Falta Produzir (Pendente)",
    f"{qtd_total_pend:,.0f} pçs".replace(",", "."),
    f"Saldo: {qtd_total_pend:,.0f} pçs".replace(",", "."),
    delta_color="inverse",
)
kpi4.metric(
    "4. Status Geral das OPs",
    f"{(df_view['STATUS'] == 'Ok').sum():,} OPs Ok".replace(",", "."),
    f"{(df_view['STATUS'] == 'Falta').sum():,} OPs em Falta".replace(",", "."),
    delta_color="off",
)

st.divider()

# --- BOTÕES DE DOWNLOAD ---
df_resumo_mensal = gerar_resumo_mensal(df_view)

col_dl_excel, col_dl_csv = st.columns([1.5, 1])
with col_dl_excel:
    excel_bytes = gerar_excel_vba(df_view, df_resumo_mensal)
    nome_arquivo_excel = f"MACRO_PRODUCAO_{datetime.now().strftime('%d-%m_%H%M')}.xlsx"
    st.download_button(
        label="📥 Baixar Pasta Excel Completa (.xlsx)",
        data=excel_bytes,
        file_name=nome_arquivo_excel,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )

with col_dl_csv:
    csv_bytes = df_view.to_csv(index=False, sep=";").encode("latin-1")
    st.download_button(
        label="📄 Baixar Visão Filtrada em CSV",
        data=csv_bytes,
        file_name="extracao_ops_filtrada.csv",
        mime="text/csv",
        use_container_width=True,
    )

# --- VISUALIZAÇÃO EM MODO CELULAR OU ABAS ---
if modo_mobile:
    st.markdown("#### 📱 Visão Otimizada para Celular")
    aba_mob_falta, aba_mob_ok = st.tabs(
        ["🔴 A Produzir (Falta)", "🟢 Produzidas (Ok)"]
    )

    with aba_mob_falta:
        df_falta_mob = df_view[df_view["STATUS"] == "Falta"]
        st.caption(f"Total: {len(df_falta_mob)} OPs pendentes")
        for _, row in df_falta_mob.head(40).iterrows():
            with st.container(border=True):
                st.markdown(
                    f"**OP:** `{row['Numero da OP']}` | **Falta:** `:red[{int(row['Saldo_Pendente']):,} pçs]`"
                )
                st.markdown(f"**Peça:** {row['Produto']} - {row['Desc. Prod.']}")
                st.caption(f"Lote/Projeto: {row['Observacao']}")

    with aba_mob_ok:
        df_ok_mob = df_view[df_view["STATUS"] == "Ok"]
        st.caption(f"Total: {len(df_ok_mob)} OPs concluídas")
        for _, row in df_ok_mob.head(40).iterrows():
            with st.container(border=True):
                st.markdown(
                    f"**OP:** `{row['Numero da OP']}` | **Qtd:** `:green[{int(row['Quantidade']):,} pçs]`"
                )
                st.markdown(f"**Peça:** {row['Produto']} - {row['Desc. Prod.']}")
                st.caption(f"Concluída em: {row['DT Real Fim']}")

else:
    # --- ABAS DE TABELAS E RELATÓRIO GRÁFICO (DESKTOP) ---
    (
        tab_pendentes,
        tab_produzidas,
        tab_todas,
        tab_resumo,
        tab_graficos,
    ) = st.tabs(
        [
            "🔴 A Produzir (Falta)",
            "🟢 Produzidas (Ok)",
            "📋 Todas as OPs Filtradas",
            "📈 Resumo Produção Mensal",
            "📊 Relatório & Gráficos",
        ]
    )

    colunas_tabela = [
        "Filial",
        "Observacao",
        "Numero da OP",
        "Produto",
        "Desc. Prod.",
        "Quantidade",
        "Qtd.Produzid",
        "Saldo_Pendente",
        "DT Real Fim",
        "STATUS",
        "MÊS-ANO",
    ]
    cols_existentes = [c for c in colunas_tabela if c in df_view.columns]

    with tab_pendentes:
        df_pendentes = df_view[df_view["STATUS"] == "Falta"]
        st.caption(f"Exibindo {len(df_pendentes):,} OPs com saldo pendente.")
        st.dataframe(
            df_pendentes[cols_existentes], use_container_width=True, height=450
        )

    with tab_produzidas:
        df_produzidas = df_view[df_view["STATUS"] == "Ok"]
        st.caption(
            f"Exibindo {len(df_produzidas):,} OPs totalmente produzidas."
        )
        st.dataframe(
            df_produzidas[cols_existentes], use_container_width=True, height=450
        )

    with tab_todas:
        st.dataframe(
            df_view[cols_existentes], use_container_width=True, height=450
        )

    with tab_resumo:
        st.dataframe(
            df_resumo_mensal.style.format(
                {
                    "Qtd_Planejada": "{:,.0f}",
                    "Total Produzido": "{:,.0f}",
                    "Saldo_Pendente": "{:,.0f}",
                }
            ),
            use_container_width=True,
            height=380,
        )

    with tab_graficos:
        st.markdown("#### 📊 Análise Visual da Produção")
        g_col1, g_col2 = st.columns([2, 1.2])

        with g_col1:
            st.markdown("**Evolução Mensal (Produzido vs. Saldo Pendente):**")
            df_graf_mes = (
                df_resumo_mensal[df_resumo_mensal["MÊS-ANO"] != "Sem Data"]
                .sort_values(by="MÊS-ANO")
                .set_index("MÊS-ANO")[["Total Produzido", "Saldo_Pendente"]]
            )
            st.bar_chart(df_graf_mes, height=320)

        with g_col2:
            st.markdown("**Status das OPs Filtradas:**")
            contagem_status = (
                df_view["STATUS"]
                .value_counts()
                .rename(index={"Ok": "Produzidas (Ok)", "Falta": "A Produzir (Falta)"})
            )
            st.dataframe(
                contagem_status.to_frame("Total de OPs"),
                use_container_width=True,
            )

        st.markdown("---")
        st.markdown("**🏆 Top 10 Projetos/Lotes com Maior Volume a Produzir:**")
        df_top_lotes = (
            df_view[df_view["STATUS"] == "Falta"]
            .groupby("Observacao")["Saldo_Pendente"]
            .sum()
            .reset_index()
            .sort_values(by="Saldo_Pendente", ascending=False)
            .head(10)
        )
        if not df_top_lotes.empty:
            df_top_lotes.columns = ["Projeto / Observação", "Peças Pendentes"]
            st.dataframe(
                df_top_lotes.style.format({"Peças Pendentes": "{:,.0f}"}),
                use_container_width=True,
            )
        else:
            st.info("Nenhuma pendência encontrada para o filtro atual.")
