import io
import os
from datetime import datetime
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def carregar_e_tratar_dados(fonte_dados):
    """Lê o CSV bruto do Protheus (SC2), pulando cabeçalhos do sistema."""
    df = pd.read_csv(
        fonte_dados,
        sep=";",
        skiprows=2,
        encoding="latin-1",
        low_memory=False,
        dtype=str,
    )

    df.columns = [c.strip() for c in df.columns]

    # Tratamento numérico
    for col in ["Quantidade", "Qtd.Produzid", "Prioridade"]:
        if col in df.columns:
            df[col] = (
                df[col]
                .astype(str)
                .str.replace(".", "", regex=False)
                .str.replace(",", ".", regex=False)
            )
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    # Cálculo de Saldo
    df["Saldo_Produzir"] = (df["Quantidade"] - df["Qtd.Produzid"]).clip(lower=0)

    # Tratamento de Datas
    for col_dt in ["DT Real Fim", "Entrega", "DT Emissao"]:
        if col_dt in df.columns:
            df[col_dt + "_Parsed"] = pd.to_datetime(
                df[col_dt], format="%d/%m/%Y", errors="coerce"
            )

    hoje = pd.to_datetime(datetime.now().strftime("%Y-%m-%d"))

    # Regras de Status da OP
    def calcular_status(row):
        qtd_plan = row.get("Quantidade", 0)
        qtd_prod = row.get("Qtd.Produzid", 0)
        dt_fim = row.get("DT Real Fim_Parsed")
        dt_entrega = row.get("Entrega_Parsed")
        situacao = str(row.get("Situacao", "")).strip().upper()

        if situacao == "SUSPENSA":
            return "OP Suspensa"
        if pd.notnull(dt_fim) or (qtd_prod >= qtd_plan and qtd_plan > 0):
            return "Encerrada"
        if qtd_prod > 0 and qtd_prod < qtd_plan:
            return "Produção Parcial"
        if pd.notnull(dt_entrega) and dt_entrega < hoje:
            return "Atrasada"
        return "Em Aberto"

    df["Status_OP"] = df.apply(calcular_status, axis=1)

    # Mês/Ano de referência (Data Real de Fim se finalizada, senão Data de Entrega)
    df["Data_Referencia"] = df["DT Real Fim_Parsed"].fillna(
        df["Entrega_Parsed"]
    )
    df["Mes_Ano"] = df["Data_Referencia"].dt.strftime("%Y-%m")

    return df


def gerar_resumo_mensal(df):
    """Agregação mensal das OPs e peças."""
    df_valido = df.dropna(subset=["Mes_Ano"]).copy()

    resumo = (
        df_valido.groupby("Mes_Ano")
        .agg(
            Total_OPs=("Numero da OP", "count"),
            Qtd_Planejada=("Quantidade", "sum"),
            Qtd_Produzida=("Qtd.Produzid", "sum"),
            Saldo_Pendente=("Saldo_Produzir", "sum"),
            OPs_Encerradas=("Status_OP", lambda s: (s == "Encerrada").sum()),
            OPs_Abertas=("Status_OP", lambda s: (s == "Em Aberto").sum()),
            OPs_Atrasadas=("Status_OP", lambda s: (s == "Atrasada").sum()),
        )
        .reset_index()
    )

    resumo["% Atingimento"] = (
        (resumo["Qtd_Produzida"] / resumo["Qtd_Planejada"].replace(0, 1)) * 100
    ).round(1)
    resumo = resumo.sort_values(by="Mes_Ano", ascending=False)
    return resumo


def gerar_excel_em_memoria(df, resumo):
    """Cria o arquivo Excel formatado diretamente na memória (BytesIO) para download instantâneo."""
    buffer = io.BytesIO()
    nome_aba_extracao = f"Extracao_{datetime.now().strftime('%d-%m_%H%M')}"

    colunas_finais = [
        "Filial",
        "Numero da OP",
        "Item",
        "Sequencia",
        "Produto",
        "Desc. Prod.",
        "Quantidade",
        "Qtd.Produzid",
        "Saldo_Produzir",
        "Status_OP",
        "Entrega",
        "DT Real Fim",
        "DT Emissao",
        "Situacao",
        "Observacao",
    ]
    colunas_presentes = [c for c in colunas_finais if c in df.columns]
    df_export = df[colunas_presentes].copy()

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df_export.to_excel(writer, sheet_name=nome_aba_extracao, index=False)
        resumo.to_excel(writer, sheet_name="RESUMO_MENSAL", index=False)

        wb = writer.book
        header_fill = PatternFill(
            start_color="1F4E78", end_color="1F4E78", fill_type="solid"
        )
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

        for sheet in wb.worksheets:
            sheet.views.sheetView[0].showGridLines = True
            for cell in sheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(
                    horizontal="center", vertical="center"
                )

            for col in sheet.columns:
                max_len = max(len(str(cell.value or "")) for cell in col[:80])
                col_letter = get_column_letter(col[0].column)
                sheet.column_dimensions[col_letter].width = max(
                    max_len + 3, 12
                )

    buffer.seek(0)
    return buffer
